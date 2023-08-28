import os
from tkinter import filedialog
import xlrd
import calendar
import csv
import openpyxl


def read_folder():
    route_folder = filedialog.askdirectory()
    file_names = os.listdir(route_folder)
    file_paths = [os.path.join(route_folder, name_file) for name_file in file_names if
                  os.path.isfile(os.path.join(route_folder, name_file))]

    return file_paths


def write_data(year, month, data_fo_f2, data_fo_f2_correct, folder: str):
    data_time = [i for i in range(24)]
    header = ['year', 'month', 'day', 'time', 'foF2', 'foF2_']
    file = folder + "/" + str(year) + " " + str(month) + ".csv"

    with open(file, 'w', newline='') as archivo_csv:
        writer = csv.writer(archivo_csv)
        writer.writerow(header)
        i = 1
        for row_fo_f2, row_fo_f2_c in zip(data_fo_f2, data_fo_f2_correct):
            for fo_f2, fo_f2_c, time in zip(row_fo_f2, row_fo_f2_c, data_time):
                # print("day: ", i, "\t time: ", time, "\tfo_f2: ", fo_f2, "\tfo_f2_correct: ", fo_f2_c)
                writer.writerow([year, month, i, time, fo_f2, fo_f2_c])
            i += 1

        archivo_csv.close()


def get_data_fo_f2_xls(worksheet, amo_days: int):
    data = []
    beg_row = 9
    end_row = amo_days + 9
    beg_col = 1
    end_col = 25
    for row in range(beg_row, end_row):
        temp = worksheet.row_values(row, beg_col, end_col)
        data.append(temp)

    return data


def get_data_fo_f2_xlsx(worksheet, amo_days: int):
    data_fo_f2 = []
    for row in worksheet.iter_rows(min_row=10, max_row=(amo_days + 9), min_col=2, max_col=25, values_only=True):
        data_fo_f2.append(row)

    return data_fo_f2


def read_and_write_xlsx(route: str, folder: str):
    try:
        workbook = openpyxl.load_workbook(route)
        worksheet_fof2 = workbook[workbook.sheetnames[0]]
        worksheet_fof2_correct = workbook[workbook.sheetnames[2]]

        month: int = int(worksheet_fof2['F6'].value)
        year: int = int(worksheet_fof2['H6'].value)
        amo_days = calendar.monthrange(year=year, month=month)

        data_fo_f2 = get_data_fo_f2_xlsx(worksheet_fof2, amo_days[1])
        data_fo_f2_correct = get_data_fo_f2_xlsx(worksheet_fof2_correct, amo_days[1])

        write_data(year, month, data_fo_f2, data_fo_f2_correct, folder)
        workbook.close()
    except IndexError:
        print("Index out range")
        bad_index.append(route)
    except Exception as e:
        bad_files.append(route)
        print(f"Other error: {e}")


def read_and_write_xls(route: str, folder: str):
    try:
        workbook = xlrd.open_workbook(route)
        worksheets_names = workbook.sheet_names()

        worksheet_fof2 = workbook.sheet_by_name(worksheets_names[0])
        worksheet_fof2_correct = workbook.sheet_by_name(worksheets_names[2])

        month: int = int(worksheet_fof2.cell_value(5, 5))
        year: int = int(worksheet_fof2.cell_value(5, 7))
        amo_days = calendar.monthrange(year=year, month=month)

        data_fo_f2 = get_data_fo_f2_xls(worksheet_fof2, amo_days[1])
        data_fo_f2_correct = get_data_fo_f2_xls(worksheet_fof2_correct, amo_days[1])

        write_data(year, month, data_fo_f2, data_fo_f2_correct, folder)
    except IndexError:
        print("Index out range")
        bad_index.append(route)
    except Exception as e:
        print(f"Other error: {e}")
        bad_files.append(route)


bad_files = []
bad_index = []


def main():
    list_routes = read_folder()
    route_folder = filedialog.askdirectory()
    print(route_folder)
    for route in list_routes:
        ext = os.path.splitext(route)[1]
        print(route)
        if '.xlsx' == ext:
            read_and_write_xlsx(route, route_folder)
        elif '.xls' == ext:
            read_and_write_xls(route, route_folder)
        else:
            print('No excel file')

    print(f"Amount Bad Files: {len(bad_files)}")
    for file in bad_files:
        print(f"\tFile: {file}")

    print(f"Amount Bad Index: {len(bad_index)}")
    for file in bad_index:
        print(f"\tFile: {file}")


if __name__ == "__main__":
    main()
