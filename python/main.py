import openpyxl
import csv
import calendar
import xlrd


def get_data_fo_f2_xlsx(worksheet, max_row):
    data_fo_f2 = []
    for row in worksheet.iter_rows(min_row=10, max_row=max_row + 9, min_col=2, max_col=25, values_only=True):
        data_fo_f2.append(row)

    return data_fo_f2


def get_data_times():
    data_times = []
    for i in range(24):
        data_times.append(i)

    return data_times


def write_data(file_out, header, year, month, data_times, data_fo_f2, data_fo_f2_correct):
    with open(file_out, 'w', newline='') as archivo_csv:
        writer = csv.writer(archivo_csv)
        writer.writerow(header)
        i = 1
        for row_fo_f2, row_fo_f2_c in zip(data_fo_f2, data_fo_f2_correct):
            for fo_f2, fo_f2_c, time in zip(row_fo_f2, row_fo_f2_c, data_times):
                print("day: ", i, "\t time: ", time, "\tfo_f2: ", fo_f2, "\tfo_f2_correct: ", fo_f2_c)
                writer.writerow([year, month, i, time, fo_f2, fo_f2_c])
            i += 1

        archivo_csv.close()

    print("file created :)")


def check_data(col_data_fo_f2, col_data_fo_f2_correct):
    for r1, r2 in zip(col_data_fo_f2, col_data_fo_f2_correct):
        if r1 != r2:
            return False

    return True


def read_and_write_xlsx(route: str):
    workbook = openpyxl.load_workbook(route)
    sheet = workbook.sheetnames[0]
    worksheet = workbook[sheet]

    data_times = get_data_times()
    month = worksheet['F6'].value
    year = worksheet['H6'].value
    amo_days = calendar.monthrange(year=year, month=month)
    data_fo_f2 = get_data_fo_f2_xlsx(worksheet, amo_days[1])

    sheet = workbook.sheetnames[2]
    worksheet = workbook[sheet]
    data_fo_f2_correct = get_data_fo_f2_xlsx(worksheet, amo_days[1])

    header = ['year', 'month', 'day', 'time', 'foF2', 'foF2_']
    file = "out1.csv"

    flag = check_data(data_fo_f2, data_fo_f2_correct)

    if flag == 1:
        write_data(file, header, year, month, data_times, data_fo_f2, data_fo_f2_correct)
    else:
        print("data bad")

    workbook.close()


def get_data_fo_f2_xls(worksheet, amo_days: int):
    data = []
    beg_row = 9
    end_row = amo_days + 9
    beg_col = 1
    end_col = 25
    for row in range(beg_row, end_row):
        temp = worksheet.row_values(row, beg_col, end_col)
        data.append(temp)

    return data


def read_and_write_xls(route: str):
    workbook = xlrd.open_workbook(route)
    print("File : " + route + " open")
    worksheets_names = workbook.sheet_names()

    worksheet1 = workbook.sheet_by_name(worksheets_names[0])
    worksheet2 = workbook.sheet_by_name(worksheets_names[2])

    month: int = int(worksheet1.cell_value(5, 5))
    year: int = int(worksheet1.cell_value(5, 7))

    amo_days = calendar.monthrange(year=year, month=month)
    data_time = get_data_times()
    header = ['year', 'month', 'day', 'time', 'foF2', 'foF2_']
    file = str(year) + " " + str(month) + ".csv"

    data_fo_f2 = get_data_fo_f2_xls(worksheet1, amo_days[1])
    data_fo_f2_correct = get_data_fo_f2_xls(worksheet2, amo_days[1])

    write_data(file, header, year, month, data_time, data_fo_f2, data_fo_f2_correct)

    print('\n')


def main():
    routes = [
        'foF2/TUCUMAN201601foF2.xls',
        'foF2/TUCUMAN201602foF2.xls',
        'foF2/TUCUMAN201603foF2.xls',
        'foF2/TUCUMAN201604foF2.xls',
        'foF2/TUCUMAN201605foF2.xls',
        'foF2/TUCUMAN201606foF2.xls',
        'foF2/TUCUMAN201607foF2.xls',
        'foF2/TUCUMAN201608foF2.xls',
        'foF2/TUCUMAN201609foF2.xls',
        'foF2/TUCUMAN201610foF2.xls',
        'foF2/TUCUMAN201611foF2.xls',
        'foF2/TUCUMAN201612foF2.xls',
    ]

    for route in routes:
        read_and_write_xls(route)


if __name__ == '__main__':
    main()
